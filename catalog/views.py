from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Book, Category, Order, OrderItem, Profile, PrintJob
import uuid
from django.contrib.auth import login
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.db.models import Q
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from datetime import timedelta
from django.db.models import Sum
from catalog.models import OrderItem
from django.http import JsonResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse


def catalog(request):
    books = Book.objects.filter(is_available=True)
    categories = Category.objects.all()

    # Фільтр за категорією
    category_slug = request.GET.get('category')
    current_category = None
    if category_slug:
        current_category = get_object_or_404(Category, slug=category_slug)
        books = books.filter(category=current_category)

    # Пошук
    query = request.GET.get('q', '').strip()
    if query:
        query_lower = query.lower()
        all_books = books
        result_ids = [
            b.pk for b in all_books
            if query_lower in b.title.lower() or query_lower in b.author.lower()
        ]
        books = Book.objects.filter(pk__in=result_ids, is_available=True)

    # Сортування
    sort = request.GET.get('sort', '')
    if sort == 'price_asc':
        books = books.order_by('price')
    elif sort == 'price_desc':
        books = books.order_by('-price')
    elif sort == 'title':
        books = books.order_by('title')
    else:
        books = books.order_by('-id')

    context = {
        'books': books,
        'categories': categories,
        'current_category': current_category,
        'query': query,
        'current_sort': sort,
    }
    return render(request, 'catalog/catalog.html', context)

def book_detail(request, pk):
    book = get_object_or_404(Book, pk=pk, is_available=True)
    return render(request, 'catalog/book_detail.html', {'book': book})

def cart(request):
    cart = request.session.get('cart', {})
    books = []
    total = 0
    for book_id, quantity in cart.items():
        try:
            book = Book.objects.get(pk=book_id)
            subtotal = book.price * quantity
            total += subtotal
            books.append({
                'book': book,
                'quantity': quantity,
                'subtotal': subtotal,
            })
        except Book.DoesNotExist:
            pass
    return render(request, 'catalog/cart.html', {'books': books, 'total': total})


def add_to_cart(request, pk):
    book = get_object_or_404(Book, pk=pk)
    cart = request.session.get('cart', {})
    key = str(pk)
    current_qty = cart.get(key, 0)

    # Перевірка наявності на складі
    if current_qty < book.stock:
        cart[key] = current_qty + 1
        request.session['cart'] = cart
    else:
        # Можна додати повідомлення що товару немає
        pass

    return redirect('cart')


def add_to_cart_ajax(request, pk):
    book = get_object_or_404(Book, pk=pk)
    cart = request.session.get('cart', {})
    key = str(pk)
    current_qty = cart.get(key, 0)

    if current_qty < book.stock:
        cart[key] = current_qty + 1
        request.session['cart'] = cart
        cart_count = sum(cart.values())
        return JsonResponse({'status': 'ok', 'cart_count': cart_count})
    else:
        return JsonResponse({'status': 'error', 'message': 'Немає в наявності'})


def remove_from_cart(request, pk):
    cart = request.session.get('cart', {})
    key = str(pk)
    if key in cart:
        if cart[key] > 1:
            cart[key] -= 1  # зменшуємо на 1
        else:
            del cart[key]   # видаляємо якщо останній
    request.session['cart'] = cart
    return redirect('cart')


def remove_from_cart_all(request, pk):
    cart = request.session.get('cart', {})
    key = str(pk)
    if key in cart:
        del cart[key]
    request.session['cart'] = cart
    return redirect('cart')


def checkout(request):
    cart = request.session.get('cart', {})

    if not cart:
        messages.warning(request, 'Кошик порожній!')
        return redirect('cart')

    def get_cart_books(cart):
        books = []
        total = 0
        for book_id, quantity in cart.items():
            try:
                book = Book.objects.get(pk=book_id)
                subtotal = book.price * quantity
                total += subtotal
                books.append({
                    'book': book,
                    'quantity': quantity,
                    'subtotal': subtotal,
                })
            except Book.DoesNotExist:
                pass
        return books, total

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        address = request.POST.get('address', '').strip()
        comment = request.POST.get('comment', '').strip()

        def render_checkout_with_error(error_msg):
            messages.error(request, error_msg)
            books, total = get_cart_books(cart)
            return render(request, 'catalog/checkout.html', {
                'books': books,
                'total': total,
                'form_data': request.POST,
            })

        if request.user.is_authenticated:
            user = request.user
        else:
            # Валідація обов'язкових полів
            if not first_name:
                return render_checkout_with_error('Ім\'я є обов\'язковим полем!')
            if not email:
                return render_checkout_with_error('Email є обов\'язковим полем!')
            if not phone:
                return render_checkout_with_error('Телефон є обов\'язковим полем!')
            if not address:
                return render_checkout_with_error('Адреса доставки є обов\'язковим полем!')

            # Валідація формату телефону
            import re
            phone_clean = re.sub(r'[\s\-\(\)]', '', phone)
            if not re.match(r'^\+?[\d]{10,13}$', phone_clean):
                return render_checkout_with_error('Невірний формат телефону! Приклад: +380991234567')

            # Валідація формату email
            if not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
                return render_checkout_with_error('Невірний формат email!')

            # Перевірка чи email вже зайнятий
            if User.objects.filter(email=email).exists():
                return render_checkout_with_error(
                    f'Користувач з email {email} вже існує. Будь ласка, увійдіть в акаунт.'
                )

            user, created = User.objects.get_or_create(
                email=email,
                defaults={
                    'username': email,
                    'first_name': first_name,
                    'last_name': last_name,
                }
            )
            if created:
                password = uuid.uuid4().hex[:10]
                user.set_password(password)
                user.save()
                login(request, user)

                # Надсилаємо пароль на email
                send_mail(
                    subject='Ваш акаунт у Книгарні',
                    message=f'Вітаємо, {first_name}!\n\nВаше замовлення оформлено. Для вас створено акаунт:\n\nЛогін: {email}\nПароль: {password}\n\nУвійдіть в особистий кабінет:\nhttp://127.0.0.1:8000/login/\n\nЗ повагою,\nКнигарня',
                    from_email='noreply@bookstore.com',
                    recipient_list=[email],
                    fail_silently=True,
                )

        # Оновлюємо профіль — телефон і адреса завжди зберігаються
        profile, _ = Profile.objects.get_or_create(user=user)
        if phone:
            profile.phone = phone
        if address:
            profile.address = address
        profile.save()

        # Створюємо замовлення
        order = Order.objects.create(
            user=user,
            comment=comment
        )

        # Додаємо позиції
        for book_id, quantity in cart.items():
            try:
                book = Book.objects.get(pk=book_id)
                OrderItem.objects.create(
                    order=order,
                    book=book,
                    quantity=quantity,
                    price=book.price
                )
                book.stock -= quantity
                book.save()
            except Book.DoesNotExist:
                pass

        request.session['cart'] = {}
        messages.success(request, f'Замовлення #{order.pk} успішно оформлено!')
        return redirect('order_detail', pk=order.pk)

    # GET — показуємо форму
    books, total = get_cart_books(cart)
    profile_phone = ''
    profile_address = ''
    if request.user.is_authenticated:
        try:
            profile_phone = request.user.profile.phone
            profile_address = request.user.profile.address
        except Profile.DoesNotExist:
            pass

    return render(request, 'catalog/checkout.html', {
        'books': books,
        'total': total,
        'profile_phone': profile_phone,
        'profile_address': profile_address,
    })



def order_detail(request, pk):
    if not request.user.is_authenticated:
        return redirect('catalog')
    order = get_object_or_404(Order, pk=pk, user=request.user)
    return render(request, 'catalog/order_detail.html', {'order': order})

@login_required
def cabinet(request):
    orders = Order.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'catalog/cabinet.html', {'orders': orders})


@login_required
def cabinet_edit(request):
    profile, _ = Profile.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        address = request.POST.get('address', '').strip()

        if not first_name:
            messages.error(request, 'Ім\'я є обов\'язковим полем!')
            return redirect('cabinet_edit')

        request.user.first_name = first_name
        request.user.last_name = last_name
        request.user.save()

        profile.phone = phone
        profile.address = address
        profile.save()

        messages.success(request, 'Профіль успішно оновлено!')
        return redirect('cabinet')

    return render(request, 'catalog/cabinet_edit.html', {'profile': profile})

from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from datetime import timedelta

@staff_member_required(login_url='login')
def admin_dashboard(request):
    import json
    today = timezone.now().date()
    month_start = today.replace(day=1)

    orders_today = Order.objects.filter(created_at__date=today).count()
    orders_total = Order.objects.count()
    orders_new = Order.objects.filter(status='new').count()

    revenue_month = sum(
        o.get_total_price()
        for o in Order.objects.filter(created_at__date__gte=month_start)
    )

    low_stock_books = Book.objects.filter(stock__lt=5, is_available=True).order_by('stock')
    recent_orders = Order.objects.select_related('user').prefetch_related('items')[:8]

    # Топ 5 книг
    from django.db.models import Sum
    top_books = (
        OrderItem.objects
        .values('book')
        .annotate(total=Sum('quantity'))
        .order_by('-total')[:5]
    )
    top_book_ids = [item['book'] for item in top_books]
    books_map = {b.pk: b for b in Book.objects.filter(pk__in=top_book_ids)}
    top_books_list = [
        {'book': books_map[item['book']], 'total': item['total']}
        for item in top_books
        if item['book'] in books_map
    ]

    # Графіки для 7, 30, 90 днів
    def build_chart_data(days, value_type='orders'):
        labels, data = [], []
        # Групуємо по тижнях якщо > 14 днів
        if days <= 14:
            for i in range(days - 1, -1, -1):
                day = today - timedelta(days=i)
                labels.append(day.strftime('%d.%m'))
                if value_type == 'orders':
                    data.append(Order.objects.filter(created_at__date=day).count())
                else:
                    orders = Order.objects.filter(created_at__date=day).prefetch_related('items__book')
                    data.append(float(sum(o.get_total_price() for o in orders)))
        else:
            # По тижнях
            weeks = days // 7
            for i in range(weeks - 1, -1, -1):
                week_end = today - timedelta(weeks=i)
                week_start = week_end - timedelta(days=6)
                labels.append(week_start.strftime('%d.%m'))
                if value_type == 'orders':
                    data.append(Order.objects.filter(
                        created_at__date__gte=week_start,
                        created_at__date__lte=week_end
                    ).count())
                else:
                    orders = Order.objects.filter(
                        created_at__date__gte=week_start,
                        created_at__date__lte=week_end
                    ).prefetch_related('items__book')
                    data.append(float(sum(o.get_total_price() for o in orders)))
        return {'labels': labels, 'data': data}

    chart_data_json = json.dumps({
        7:  build_chart_data(7, 'orders'),
        30: build_chart_data(30, 'orders'),
        90: build_chart_data(90, 'orders'),
    })

    revenue_data_json = json.dumps({
        7:  build_chart_data(7, 'revenue'),
        30: build_chart_data(30, 'revenue'),
        90: build_chart_data(90, 'revenue'),
    })

    context = {
        'orders_today': orders_today,
        'orders_total': orders_total,
        'orders_new': orders_new,
        'revenue_month': revenue_month,
        'low_stock_books': low_stock_books,
        'recent_orders': recent_orders,
        'top_books': top_books_list,
        'chart_data_json': chart_data_json,
        'revenue_data_json': revenue_data_json,
    }
    return render(request, 'catalog/admin/dashboard.html', context)


@staff_member_required(login_url='login')
def admin_orders(request):
    orders = Order.objects.select_related('user').prefetch_related('items').order_by('-created_at')

    status_filter = request.GET.get('status', '')
    if status_filter:
        orders = orders.filter(status=status_filter)

    context = {
        'orders': orders,
        'status_filter': status_filter,
        'status_choices': Order.STATUS_CHOICES,
    }
    return render(request, 'catalog/admin/orders.html', context)


@staff_member_required(login_url='login')
def admin_order_detail(request, pk):
    order = get_object_or_404(Order, pk=pk)
    print_job, _ = PrintJob.objects.get_or_create(order=order)

    if request.method == 'POST':
        new_order_status = request.POST.get('status')
        if new_order_status:
            order.status = new_order_status
            order.save()

        new_print_status = request.POST.get('print_status', print_job.status)

        # Автоматична фіксація часових міток
        if new_print_status == 'printing' and print_job.status != 'printing':
            if not print_job.started_at:
                print_job.started_at = timezone.now()

        if new_print_status == 'done' and print_job.status != 'done':
            if not print_job.finished_at:
                print_job.finished_at = timezone.now()

        print_job.status = new_print_status
        print_job.notes = request.POST.get('notes', print_job.notes)

        # Технологічні параметри
        print_job.format = request.POST.get('format', print_job.format)
        circulation = request.POST.get('circulation', '')
        if circulation:
            print_job.circulation = int(circulation)
        print_job.print_method = request.POST.get('print_method', print_job.print_method)
        print_job.paper_type = request.POST.get('paper_type', print_job.paper_type)
        print_job.binding_type = request.POST.get('binding_type', print_job.binding_type)

        print_job.save()
        messages.success(request, 'Замовлення оновлено.')
        return redirect('admin_order_detail', pk=pk)

    context = {
        'order': order,
        'print_job': print_job,
        'status_choices': Order.STATUS_CHOICES,
        'print_status_choices': PrintJob.STATUS_CHOICES,
        'print_method_choices': PrintJob.PRINT_METHOD_CHOICES,
        'paper_choices': PrintJob.PAPER_CHOICES,
        'binding_choices': PrintJob.BINDING_CHOICES,
    }
    return render(request, 'catalog/admin/order_detail.html', context)


@staff_member_required(login_url='login')
def admin_order_status(request, pk):
    """Швидка зміна статусу прямо з таблиці замовлень"""
    if request.method == 'POST':
        order = get_object_or_404(Order, pk=pk)
        new_status = request.POST.get('status')
        if new_status:
            order.status = new_status
            order.save()
    return redirect(request.POST.get('next', 'admin_orders'))


@staff_member_required(login_url='login')
def admin_books(request):
    books = Book.objects.select_related('category').order_by('-created_at')

    q = request.GET.get('q', '').strip()
    if q:
        books = [b for b in books if q.lower() in b.title.lower() or q.lower() in b.author.lower()]

    context = {
        'books': books,
        'q': q,
    }
    return render(request, 'catalog/admin/books.html', context)


@staff_member_required(login_url='login')
def admin_book_add(request):
    categories = Category.objects.all()
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        author = request.POST.get('author', '').strip()
        price = request.POST.get('price')
        stock = request.POST.get('stock', 0)
        description = request.POST.get('description', '')
        category_id = request.POST.get('category')
        is_available = request.POST.get('is_available') == 'on'
        cover = request.FILES.get('cover')

        if not title or not author or not price:
            messages.error(request, 'Заповніть обов\'язкові поля.')
        else:
            book = Book.objects.create(
                title=title,
                author=author,
                price=price,
                stock=stock,
                description=description,
                category_id=category_id or None,
                is_available=is_available,
                cover=cover,
            )
            messages.success(request, f'Книгу «{book.title}» додано.')
            return redirect('admin_books')

    return render(request, 'catalog/admin/book_form.html', {'categories': categories, 'book': None})


@staff_member_required(login_url='login')
def admin_book_edit(request, pk):
    book = get_object_or_404(Book, pk=pk)
    categories = Category.objects.all()

    if request.method == 'POST':
        book.title = request.POST.get('title', book.title).strip()
        book.author = request.POST.get('author', book.author).strip()
        book.price = request.POST.get('price', book.price)
        book.stock = request.POST.get('stock', book.stock)
        book.description = request.POST.get('description', book.description)
        category_id = request.POST.get('category')
        book.category_id = category_id or None
        book.is_available = request.POST.get('is_available') == 'on'
        if request.FILES.get('cover'):
            book.cover = request.FILES['cover']
        book.save()
        messages.success(request, f'Книгу «{book.title}» оновлено.')
        return redirect('admin_books')

    return render(request, 'catalog/admin/book_form.html', {'categories': categories, 'book': book})


@staff_member_required(login_url='login')
def admin_book_delete(request, pk):
    book = get_object_or_404(Book, pk=pk)
    if request.method == 'POST':
        title = book.title
        book.delete()
        messages.success(request, f'Книгу «{title}» видалено.')
    return redirect('admin_books')


@staff_member_required(login_url='login')
def admin_users(request):
    from django.contrib.auth.models import User
    users = User.objects.prefetch_related('orders').order_by('-date_joined')
    context = {'users': users}
    return render(request, 'catalog/admin/users.html', context)


def admin_stats(request):
    if request.user.is_authenticated and request.user.is_staff:
        return {'orders_new_count': Order.objects.filter(status='new').count()}
    return {}


@staff_member_required(login_url='login')
def admin_export_orders(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Замовлення'

    # Стилі
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A1814', end_color='1A1814', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    alt_fill = PatternFill(start_color='F7F4EE', end_color='F7F4EE', fill_type='solid')

    thin = Side(style='thin', color='E6DDD0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    status_map = {
        'new': 'Нове',
        'confirmed': 'Підтверджено',
        'printing': 'В друці',
        'ready': 'Готово',
        'shipped': 'Відправлено',
        'delivered': 'Доставлено',
    }

    print_status_map = {
        'pending': 'Очікує',
        'prepress': 'Додрукарська підготовка',
        'printing': 'Друк',
        'postpress': 'Післядрукарська обробка',
        'done': 'Виконано',
    }

    # Заголовки
    headers = [
        '№', 'Дата', 'Покупець', 'Email', 'Телефон',
        'Адреса', 'Книги', 'Кількість', 'Сума (грн)',
        'Статус замовлення', 'Статус друку', 'Коментар'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 22

    # Фільтр по статусу якщо є
    status_filter = request.GET.get('status', '')
    orders = Order.objects.select_related('user__profile').prefetch_related('items__book').order_by('-created_at')
    if status_filter:
        orders = orders.filter(status=status_filter)

    # Дані
    for row_idx, order in enumerate(orders, 2):
        is_alt = row_idx % 2 == 0

        books_titles = '\n'.join([
            f'{item.book.title} ({item.book.author})'
            for item in order.items.all()
        ])
        books_qty = '\n'.join([
            str(item.quantity)
            for item in order.items.all()
        ])

        try:
            phone = order.user.profile.phone
            address = order.user.profile.address
        except Exception:
            phone = ''
            address = ''

        try:
            print_status = print_status_map.get(order.print_job.status, '—')
        except Exception:
            print_status = '—'

        row_data = [
            order.pk,
            order.created_at.strftime('%d.%m.%Y %H:%M'),
            order.user.get_full_name() or order.user.username,
            order.user.email,
            phone,
            address,
            books_titles,
            books_qty,
            float(order.get_total_price()),
            status_map.get(order.status, order.status),
            print_status,
            order.comment,
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if is_alt:
                cell.fill = alt_fill

        ws.row_dimensions[row_idx].height = max(
            15 * len(order.items.all()), 18
        )

    # Ширина колонок
    col_widths = [6, 18, 22, 28, 16, 35, 45, 12, 14, 20, 14, 30]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = width

    # Закріпити перший рядок
    ws.freeze_panes = 'A2'

    # Другий лист — статистика
    ws2 = wb.create_sheet(title='Статистика')

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 18

    stat_header_font = Font(bold=True, color='FFFFFF', size=11)
    stat_header_fill = PatternFill(start_color='B85C38', end_color='B85C38', fill_type='solid')

    stat_title = ws2.cell(row=1, column=1, value='Показник')
    stat_title.font = stat_header_font
    stat_title.fill = stat_header_fill
    stat_title.border = border
    stat_title.alignment = header_align

    stat_val = ws2.cell(row=1, column=2, value='Значення')
    stat_val.font = stat_header_font
    stat_val.fill = stat_header_fill
    stat_val.border = border
    stat_val.alignment = header_align

    # Статистика
    from django.utils import timezone
    today = timezone.now().date()
    month_start = today.replace(day=1)

    all_orders = Order.objects.prefetch_related('items')
    revenue_total = sum(o.get_total_price() for o in all_orders)
    revenue_month = sum(
        o.get_total_price()
        for o in all_orders.filter(created_at__date__gte=month_start)
    )

    stats = [
        ('Всього замовлень', Order.objects.count()),
        ('Нових замовлень', Order.objects.filter(status='new').count()),
        ('Підтверджених', Order.objects.filter(status='confirmed').count()),
        ('В друці', Order.objects.filter(status='printing').count()),
        ('Готово', Order.objects.filter(status='ready').count()),
        ('Відправлено', Order.objects.filter(status='shipped').count()),
        ('Доставлено', Order.objects.filter(status='delivered').count()),
        ('', ''),
        ('Виручка всього (грн)', float(revenue_total)),
        ('Виручка за місяць (грн)', float(revenue_month)),
        ('Замовлень сьогодні', Order.objects.filter(created_at__date=today).count()),
        ('', ''),
        ('Книг в каталозі', Book.objects.filter(is_available=True).count()),
        ('Користувачів', User.objects.count()),
        ('Дата експорту', today.strftime('%d.%m.%Y')),
    ]

    for row_idx, (label, value) in enumerate(stats, 2):
        c1 = ws2.cell(row=row_idx, column=1, value=label)
        c2 = ws2.cell(row=row_idx, column=2, value=value)
        if label:
            c1.border = border
            c2.border = border
            if row_idx % 2 == 0:
                c1.fill = alt_fill
                c2.fill = alt_fill
            c2.alignment = Alignment(horizontal='right')

    # Відповідь
    filename = f'orders_{today.strftime("%d-%m-%Y")}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response